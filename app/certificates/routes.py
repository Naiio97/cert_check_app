from flask import Blueprint, request, jsonify, current_app, flash, redirect, url_for, render_template, send_from_directory, g
from app.models import Certifikat, Server, AuditLog
from app import db
from sqlalchemy import text
from app.utils import allowed_file, is_valid_date
from datetime import datetime, timedelta
import pandas as pd
import os
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from app import get_expiry_class  # Přidejte tento import na začátek souboru
from tempfile import NamedTemporaryFile
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

bp = Blueprint('certificates', __name__)

@bp.route('/pridat', methods=['GET', 'POST'])
def pridat_certifikat():
    if request.method == 'POST':
        try:
            expirace = datetime.strptime(request.form['expirace'], '%d.%m.%Y')
            
            novy_cert = Certifikat(
                server=request.form['server'],
                cesta=request.form['cesta'],
                nazev=request.form['nazev'],
                expirace=expirace,
                poznamka=request.form.get('poznamka', '')
            )
            db.session.add(novy_cert)
            db.session.flush()
            db.session.add(AuditLog(
                akce='pridano',
                certifikat_nazev=novy_cert.nazev,
                server=novy_cert.server,
                detail=f'Cesta: {novy_cert.cesta}, Expirace: {novy_cert.expirace.strftime("%d.%m.%Y")}'
            ))
            db.session.commit()
            flash('Certifikát byl úspěšně přidán!')
            return redirect(url_for('main.index'))
        except ValueError:
            db.session.rollback()
            flash('Neplatný formát data! Použijte formát dd.mm.yyyy (např. 31.12.2025)', 'error')
            servery = Server.query.all()
            return render_template('formular.html', certifikat=None, servery=servery, error_field='expirace')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'Chyba při přidávání certifikátu: {str(e)}')
            flash(f'Chyba při přidávání certifikátu: {str(e)}')
            servery = Server.query.all()
            return render_template('formular.html', certifikat=None, servery=servery)
    
    servery = Server.query.all()
    return render_template('formular.html', certifikat=None, servery=servery)

@bp.route('/get-edit-form/<int:id>')
def get_edit_form(id):
    try:
        certifikat = Certifikat.query.get_or_404(id)
        servery = Server.query.all()
        return render_template('edit_modal.html', certifikat=certifikat, servery=servery)
    except Exception as e:
        current_app.logger.error(f'Chyba při načítání editačního formuláře: {str(e)}')
        return jsonify({'error': str(e)}), 500

@bp.route('/upravit/<int:id>', methods=['GET', 'POST'])
def upravit_certifikat(id):
    certifikat = Certifikat.query.get_or_404(id)
    servery = Server.query.all()
    
    if request.method == 'POST':
        try:
            certifikat.server = request.form['server']
            certifikat.cesta = request.form['cesta']
            certifikat.nazev = request.form['nazev']
            
            # Zpracování data
            expirace_str = request.form['expirace']
            expirace = datetime.strptime(expirace_str, '%d.%m.%Y').date()
            certifikat.expirace = expirace
            
            certifikat.poznamka = request.form.get('poznamka', '')
            
            db.session.add(AuditLog(
                akce='upraveno',
                certifikat_nazev=certifikat.nazev,
                server=certifikat.server,
                detail=f'Cesta: {certifikat.cesta}, Expirace: {expirace.strftime("%d.%m.%Y")}'
            ))
            db.session.commit()
            flash('Certifikát byl úspěšně upraven', 'success')
            return redirect('/evidence_certifikatu')
        except ValueError:
            db.session.rollback()
            flash('Neplatný formát data! Použijte formát dd.mm.yyyy', 'error')
            return render_template('formular.html', certifikat=certifikat, servery=servery, error_field='expirace')
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'Chyba při úpravě certifikátu: {str(e)}')
            flash(f'Chyba při úpravě certifikátu: {str(e)}', 'error')
            return render_template('formular.html', certifikat=certifikat, servery=servery)
    
    return render_template('formular.html', 
                         certifikat=certifikat,
                         servery=servery)

@bp.route('/smazat/<int:id>', methods=['POST'])
def smazat_certifikat(id):
    try:
        certifikat = Certifikat.query.get_or_404(id)
        current_app.logger.info(f'Mazání certifikátu: {certifikat.nazev} ze serveru {certifikat.server}')
        db.session.add(AuditLog(
            akce='smazano',
            certifikat_nazev=certifikat.nazev,
            server=certifikat.server,
            detail=f'Cesta: {certifikat.cesta}, Expirace: {certifikat.expirace.strftime("%d.%m.%Y")}'
        ))
        db.session.delete(certifikat)
        db.session.commit()
        flash('Certifikát byl smazán!')
    except Exception as e:
        current_app.logger.error(f'Chyba při mazání certifikátu: {str(e)}')
        flash(f'Chyba při mazání certifikátu: {str(e)}')
    return redirect(url_for('main.index'))

@bp.route('/import-excel', methods=['POST'])
def import_excel():
    try:
        current_app.logger.info('Začátek importu z Excelu')
        if 'excel_file' not in request.files:
            current_app.logger.warning('Nebyl vybrán žádný soubor')
            return jsonify({'error': 'Nebyl vybrán žádný soubor'}), 400
        
        file = request.files['excel_file']
        if file.filename == '':
            current_app.logger.warning('Prázdný název souboru')
            return jsonify({'error': 'Nebyl vybrán žádný soubor'}), 400
        
        if not allowed_file(file.filename):
            current_app.logger.warning(f'Nepovolený typ souboru: {file.filename}')
            return jsonify({'error': 'Nepovolený typ souboru'}), 400
        
        filename = secure_filename(file.filename)
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        df = pd.read_excel(filepath)
        
        # Zpracování dat z Excelu
        df['Server'] = df['Server'].ffill()
        df['Cesta'] = df['Cesta'].ffill()
        df['Název certifikátu'] = df['Název certifikátu'].fillna('Neznámý certifikát')
        
        df = df.dropna(subset=['Server'])
        df = df[df['Expirace'].apply(is_valid_date)]
        
        # Import serverů (pouze do aktivního prostředí)
        unique_servers = df['Server'].unique()
        for server_name in unique_servers:
            if pd.isna(server_name):
                continue
            if not Server.query.filter_by(nazev=server_name).first():
                new_server = Server(nazev=server_name)
                db.session.add(new_server)
        db.session.commit()
        
        # Import certifikátů
        pridano = aktualizovano = beze_zmeny = 0
        
        for _, row in df.iterrows():
            expirace = datetime.strptime(row['Expirace'], '%d.%m.%Y').date() if isinstance(row['Expirace'], str) else row['Expirace'].date()
            
            existing = Certifikat.query.filter_by(
                server=row['Server'],
                cesta=row['Cesta'],
                nazev=row['Název certifikátu']
            ).first()
            
            if existing:
                if existing.expirace != expirace:
                    existing.expirace = expirace
                    aktualizovano += 1
                else:
                    beze_zmeny += 1
            else:
                new_cert = Certifikat(
                    server=row['Server'],
                    cesta=row['Cesta'],
                    nazev=row['Název certifikátu'],
                    expirace=expirace
                )
                db.session.add(new_cert)
                pridano += 1
                
        db.session.commit()
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'message': f'Import dokončen. Přidáno: {pridano}, Aktualizováno: {aktualizovano}'
        })
        
    except Exception as e:
        current_app.logger.error(f'Chyba při importu: {str(e)}', exc_info=True)
        return jsonify({
            'success': False,
            'message': f'Chyba při importu: {str(e)}'
        }), 400

@bp.route('/export-excel')
def export_excel():
    try:
        certifikaty = Certifikat.query.order_by(Certifikat.server, Certifikat.cesta).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Certifikáty"
        
        # Styly
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        alt_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        expired_fill = PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Hlavička
        headers = ["Server", "Cesta", "Název certifikátu", "Expirace"]
        ws.append(headers)
        
        # Formátování hlavičky
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Data
        today = datetime.now().date()
        row_num = 2
        merge_ranges = {'server': [], 'path': []}
        prev_server = None
        prev_path = None
        merge_start = {'server': 2, 'path': 2}
        
        for cert in certifikaty:
            row = [
                cert.server,
                cert.cesta,
                cert.nazev,
                cert.expirace.strftime('%d.%m.%Y')
            ]
            ws.append(row)
            
            # Formátování řádku
            for cell in ws[row_num]:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Alternující barvy řádků
                if row_num % 2 == 0:
                    cell.fill = alt_fill
                
                # Červené podbarvení pro expirované certifikáty - celý řádek
                if cert.expirace < today:
                    cell.fill = expired_fill
            
            # Kontrola pro sloučení buněk serveru
            if prev_server != cert.server and prev_server is not None:
                if row_num - merge_start['server'] > 1:
                    merge_ranges['server'].append(f'A{merge_start["server"]}:A{row_num-1}')
                merge_start['server'] = row_num
            
            # Kontrola pro sloučení buněk cesty
            if prev_path != cert.cesta or prev_server != cert.server:
                if row_num - merge_start['path'] > 1:
                    merge_ranges['path'].append(f'B{merge_start["path"]}:B{row_num-1}')
                merge_start['path'] = row_num
            
            prev_server = cert.server
            prev_path = cert.cesta
            row_num += 1
        
        # Poslední sloučení pro server a cestu
        if row_num - merge_start['server'] > 1:
            merge_ranges['server'].append(f'A{merge_start["server"]}:A{row_num-1}')
        if row_num - merge_start['path'] > 1:
            merge_ranges['path'].append(f'B{merge_start["path"]}:B{row_num-1}')
        
        # Provedení sloučení buněk
        for ranges in merge_ranges.values():
            for cell_range in ranges:
                ws.merge_cells(cell_range)
        
        # Automatická šířka sloupců
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Zamrazení hlavičky
        ws.freeze_panes = "A2"
        
        # Export
        with NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            filename = os.path.basename(tmp.name)
            wb.save(tmp.name)
            
            return send_from_directory(
                directory=os.path.dirname(tmp.name),
                path=filename,
                as_attachment=True,
                download_name=f"certifikaty_export_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
    except Exception as e:
        current_app.logger.error(f'Chyba při exportu do Excelu: {str(e)}')
        flash(f'Chyba při exportu do Excelu: {str(e)}', 'error')
        return redirect(url_for('main.index'))

@bp.route('/smazat-vse', methods=['POST'])
def smazat_vse():
    try:
        # Smazat pouze v aktivním prostředí (live/test)
        env = getattr(g, 'db_bind', 'live')
        try:
            engine = db.engines[env]
        except Exception:
            engine = db.get_engine(current_app, bind=env)
        with engine.begin() as conn:
            conn.execute(text('DELETE FROM certifikat'))
            conn.execute(text('DELETE FROM server'))
        db.session.add(AuditLog(
            akce='smazano',
            certifikat_nazev='VŠE',
            server='VŠE',
            detail=f'Smazána celá databáze ({env.upper()})'
        ))
        db.session.commit()
        flash(f'Data byla smazána v prostředí: {env.upper()}!')
    except Exception as e:
        current_app.logger.error(f'Chyba při mazání ({env}): {str(e)}')
        flash(f'Chyba při mazání: {str(e)}', 'error')
    return redirect(url_for('main.index'))

@bp.route('/detail/<int:id>')
def detail_certifikatu(id):
    try:
        certifikat = Certifikat.query.get_or_404(id)
        return render_template('detail_modal.html', certifikat=certifikat)
    except Exception as e:
        current_app.logger.error(f'Chyba při načítání detailu: {str(e)}')
        return f'Chyba při načítání detailu: {str(e)}', 500

@bp.route('/server/<server>')
def get_certifikaty_server(server):
    try:
        page = request.args.get('page', 1, type=int)
        
        pagination = Certifikat.query.filter_by(server=server)\
            .order_by(Certifikat.expirace, Certifikat.cesta)\
            .paginate(page=page, per_page=25, error_out=False)
        
        certifikaty = pagination.items
        
        certs_data = [{
            'id': cert.id,
            'server': cert.server,
            'cesta': cert.cesta,
            'nazev': cert.nazev,
            'expirace': cert.expirace.strftime('%d.%m.%Y'),
            'expiry_class': get_expiry_class(cert)
        } for cert in certifikaty]
        
        return jsonify({
            'items': certs_data,
            'pagination': {
                'page': pagination.page,
                'pages': pagination.pages,
                'total': pagination.total,
                'has_next': pagination.has_next,
                'has_prev': pagination.has_prev,
                'next_num': pagination.next_num,
                'prev_num': pagination.prev_num
            }
        })
    except Exception as e:
        current_app.logger.error(f'Chyba při načítání certifikátů: {str(e)}')
        return jsonify({'error': str(e)}), 500

@bp.route('/send-report', methods=['POST'])
def trigger_report():
    try:
        from app.tasks import send_monthly_certificate_report
        env = getattr(g, 'db_bind', 'live')
        send_monthly_certificate_report(env)
        return jsonify({'success': True, 'message': f'Report ({env}) odeslán'})
    except Exception as e:
        current_app.logger.error(f"Chyba při odesílání reportu: {str(e)}")
        return jsonify({'success': False, 'message': f'Chyba: {str(e)}'}), 500


@bp.route('/smazat-vybrane', methods=['POST'])
def smazat_vybrane():
    """Bulk delete: accepts JSON { ids: [1, 2, 3] }"""
    try:
        data = request.get_json()
        ids = data.get('ids', []) if data else []
        if not ids:
            return jsonify({'success': False, 'message': 'Žádné certifikáty k smazání'}), 400

        count = 0
        for cert_id in ids:
            cert = Certifikat.query.get(cert_id)
            if cert:
                db.session.add(AuditLog(
                    akce='smazano',
                    certifikat_nazev=cert.nazev,
                    server=cert.server,
                    detail=f'Hromadné smazání – Cesta: {cert.cesta}'
                ))
                db.session.delete(cert)
                count += 1
        db.session.commit()
        return jsonify({'success': True, 'message': f'Smazáno {count} certifikátů'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Chyba při hromadném mazání: {str(e)}')
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/export-vybrane', methods=['POST'])
def export_vybrane():
    """Export selected certificates to Excel."""
    try:
        data = request.get_json()
        ids = data.get('ids', []) if data else []
        if not ids:
            return jsonify({'success': False, 'message': 'Žádné certifikáty k exportu'}), 400

        certs = Certifikat.query.filter(Certifikat.id.in_(ids)).order_by(Certifikat.expirace).all()
        if not certs:
            return jsonify({'success': False, 'message': 'Certifikáty nenalezeny'}), 404

        from openpyxl import Workbook
        from tempfile import NamedTemporaryFile

        wb = Workbook()
        ws = wb.active
        ws.title = 'Vybrané certifikáty'
        ws.append(['Server', 'Cesta', 'Název', 'Expirace', 'Poznámka'])

        for cert in certs:
            ws.append([
                cert.server,
                cert.cesta,
                cert.nazev,
                cert.expirace.strftime('%d.%m.%Y'),
                cert.poznamka or ''
            ])

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

        tmp = NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(tmp.name)
        tmp.close()

        import os
        return send_from_directory(
            os.path.dirname(tmp.name),
            os.path.basename(tmp.name),
            as_attachment=True,
            download_name='vybrane_certifikaty.xlsx'
        )
    except Exception as e:
        current_app.logger.error(f'Chyba při exportu vybraných: {str(e)}')
        return jsonify({'success': False, 'message': str(e)}), 500